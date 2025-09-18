# app/services/wbs_generator.py
import io
import re
from typing import Tuple, Dict

import pandas as pd
from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _money(val):
    try:
        return float(val or 0.0)
    except Exception:
        return 0.0


def _find_wbs_header(ws: Worksheet) -> Tuple[int, Dict[str, int]]:
    """
    Robustly locate the header row on WEEKLY and build a column index map.
    Accepts common variants (merged cells, 'Totals' vs 'Total', etc.).
    """
    req_aliases = {
        "ssn":        ["ssn"],
        "name":       ["employee name", "employee", "name"],
        "status":     ["status"],
        "type":       ["type"],
        "rate":       ["pay rate", "payrate", "rate"],
        "dept":       ["dept", "department"],
        "a01":        ["a01", "regular", "reg"],
        "a02":        ["a02", "overtime", "ot"],
        "a03":        ["a03", "doubletime", "dt"],
    }
    opt_aliases = {
        "reg_amt":     ["a01 $", "a01$", "reg $", "regular $", "regular amt", "a01 amount"],
        "ot_amt":      ["a02 $", "a02$", "ot $",  "overtime $", "overtime amt", "a02 amount"],
        "dt_amt":      ["a03 $", "a03$", "dt $",  "doubletime $", "doubletime amt", "a03 amount"],
        "total_amt":   ["total $", "total$", "grand total $"],
        "total_plain": ["total", "totals"],
    }

    def norm_cell(v):
        v = "" if v is None else str(v)
        v = v.replace("\n", " ").replace("\r", " ")
        v = re.sub(r"\s+", " ", v).strip().lower()
        return v

    best_row, best_score, best_map = None, -1, None
    for r in range(1, ws.max_row + 1):
        row_vals = [norm_cell(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if sum(1 for v in row_vals if v) < 3:
            continue

        score = 0
        for aliases in req_aliases.values():
            if any(any(a in v for a in aliases) for v in row_vals):
                score += 1

        if score > best_score:
            col_map = {}
            for c, v in enumerate(row_vals, start=1):
                if v and v not in col_map:
                    col_map[v] = c

            def pick(aliases):
                for a in aliases:
                    if a in col_map:
                        return col_map[a]
                    for v, c in col_map.items():
                        if a in v:
                            return c
                return None

            mapped = {}
            for k, aliases in req_aliases.items():
                mapped[k] = pick(aliases)

            must_ok = (mapped.get("name") is not None and
                       mapped.get("a01") is not None and
                       mapped.get("a02") is not None and
                       mapped.get("a03") is not None)

            if must_ok:
                mapped["reg_amt"]      = pick(opt_aliases["reg_amt"])
                mapped["ot_amt"]       = pick(opt_aliases["ot_amt"])
                mapped["dt_amt"]       = pick(opt_aliases["dt_amt"])
                mapped["total_amt"]    = pick(opt_aliases["total_amt"])
                mapped["total_plain"]  = pick(opt_aliases["total_plain"])

                best_row, best_score, best_map = r, score, mapped

    if best_row is None or best_map is None:
        raise HTTPException(422, "WEEKLY header not found. Expect columns like 'Employee Name', 'A01', 'A02', 'A03'.")
    return best_row, best_map


def write_into_wbs_template(template_bytes: bytes, weekly: pd.DataFrame) -> bytes:
    """
    Writes ONLY to 'WEEKLY' sheet to match the WBS layout precisely.
    Populates SSN, Employee Name, Status, Type, Pay Rate, Dept, A01, A02, A03,
    and amount columns when present; otherwise uses far-right Total/Totals.
    """
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb["WEEKLY"] if "WEEKLY" in wb.sheetnames else wb.active

    header_row, cols = _find_wbs_header(ws)
    first_data_row = header_row + 1

    # Clear existing data rows (keep styling)
    scan_col = cols.get("name") or cols.get("ssn") or 2
    last = ws.max_row
    last_data = first_data_row - 1
    for r in range(first_data_row, last + 1):
        if ws.cell(r, scan_col).value not in (None, ""):
            last_data = r
    if last_data >= first_data_row:
        ws.delete_rows(first_data_row, last_data - first_data_row + 1)

    # Append rows
    for _, row in weekly.iterrows():
        values = [""] * max(ws.max_column, 64)
        if cols.get("ssn"):        values[cols["ssn"] - 1]         = row.get("ssn", "")
        if cols.get("name"):       values[cols["name"] - 1]        = row.get("employee", "")
        if cols.get("status"):     values[cols["status"] - 1]      = row.get("Status", "A")
        if cols.get("type"):       values[cols["type"] - 1]        = row.get("Type", "H")
        if cols.get("rate"):       values[cols["rate"] - 1]        = row.get("rate", 0.0)
        if cols.get("dept"):       values[cols["dept"] - 1]        = row.get("department", "")
        if cols.get("a01"):        values[cols["a01"] - 1]         = row.get("REG", 0.0)
        if cols.get("a02"):        values[cols["a02"] - 1]         = row.get("OT", 0.0)
        if cols.get("a03"):        values[cols["a03"] - 1]         = row.get("DT", 0.0)

        if cols.get("reg_amt"):    values[cols["reg_amt"] - 1]     = row.get("REG_$", 0.0)
        if cols.get("ot_amt"):     values[cols["ot_amt"] - 1]      = row.get("OT_$", 0.0)
        if cols.get("dt_amt"):     values[cols["dt_amt"] - 1]      = row.get("DT_$", 0.0)
        if cols.get("total_amt"):  values[cols["total_amt"] - 1]   = row.get("TOTAL_$", 0.0)
        elif cols.get("total_plain"):
            values[cols["total_plain"] - 1] = row.get("TOTAL_$", 0.0)

        ws.append(values)

    # Spacer + TOTAL row
    ws.append([])
    totals = {
        "REG":     float(weekly["REG"].sum()),
        "OT":      float(weekly["OT"].sum()),
        "DT":      float(weekly["DT"].sum()),
        "REG_$":   float(weekly["REG_$"].sum()),
        "OT_$":    float(weekly["OT_$"].sum()),
        "DT_$":    float(weekly["DT_$"].sum()),
        "TOTAL_$": float(weekly["TOTAL_$"].sum()),
    }
    row_vals = [""] * max(ws.max_column, 64)
    if cols.get("name"):       row_vals[cols["name"] - 1]        = "TOTAL"
    if cols.get("a01"):        row_vals[cols["a01"] - 1]         = _money(totals["REG"])
    if cols.get("a02"):        row_vals[cols["a02"] - 1]         = _money(totals["OT"])
    if cols.get("a03"):        row_vals[cols["a03"] - 1]         = _money(totals["DT"])
    if cols.get("reg_amt"):    row_vals[cols["reg_amt"] - 1]     = _money(totals["REG_$"])
    if cols.get("ot_amt"):     row_vals[cols["ot_amt"] - 1]      = _money(totals["OT_$"])
    if cols.get("dt_amt"):     row_vals[cols["dt_amt"] - 1]      = _money(totals["DT_$"])
    if cols.get("total_amt"):  row_vals[cols["total_amt"] - 1]   = _money(totals["TOTAL_$"])
    elif cols.get("total_plain"):
        row_vals[cols["total_plain"] - 1] = _money(totals["TOTAL_$"])
    ws.append(row_vals)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()
