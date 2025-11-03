# app/services/excel_processor.py
# Sierra → WBS translator (FULLY CALCULATED; SUPPORTS "Mon/Tue/..." TEXT DAYS)
# - Reads Sierra daily logs (Days, Name, Hours, Rate)
# - Accepts Days as real dates OR weekday text ("Mon", "Tue", ...), maps to Mon=0..Sun=6
# - Splits hours into REG (≤8/day), OT (8–12/day), DT (>12/day) + overlays WEEKLY >40 into OT
# - Computes Mon–Fri per-day hours and per-day totals (HRS × Pay Rate)
# - Fills Status/Type/Dept from roster (optional); safe defaults if missing
# - Writes blanks instead of 0s; injects Totals formula; preserves WBS name order then A–Z by last name

from __future__ import annotations

import io, re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional, Tuple, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

# ---------- Paths ----------
HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent.parent
REPO_ROOT = PROJECT_ROOT.parent

WBS_TEMPLATE_PATH = REPO_ROOT / "wbs_template.xlsx"
ROSTER_PATH       = REPO_ROOT / "roster.xlsx"

# ---------- helpers ----------
def _num(s) -> Optional[float]:
    if s is None: return None
    if isinstance(s,(int,float)):
        try: return float(s)
        except: return None
    ss = str(s).strip().replace("$","").replace(",","")
    m = re.search(r"-?\d+(\.\d+)?", ss)
    return float(m.group(0)) if m else None

def _safe_int(x) -> Optional[int]:
    try:
        if pd.isna(x): return None
    except Exception:
        pass
    try: return int(float(str(x).replace(",","").strip()))
    except Exception: return None

def _normalize_name_for_join(name: str) -> Tuple[str,str]:
    if not isinstance(name,str): return ("","")
    s = " ".join(name.replace(","," ").split()).strip()
    if not s: return ("","")
    parts = s.split(" ")
    if "," in name:
        last, rest = [x.strip() for x in name.split(",",1)]
        first = rest.split(" ")[0] if rest else ""
    else:
        first, last = parts[0], parts[-1]
    ln = "".join(ch for ch in last.lower()  if ch.isalpha())
    fn = "".join(ch for ch in first.lower() if ch.isalpha())
    return ln, fn

def _best_join(left_df: pd.DataFrame, right_df: pd.DataFrame,
               left_name_col: str, right_name_col: str) -> pd.DataFrame:
    L = left_df.copy(); R = right_df.copy()
    L[["__ln","__fn"]] = L[left_name_col].apply(lambda s: pd.Series(_normalize_name_for_join(str(s))))
    R[["__ln","__fn"]] = R[right_name_col].apply(lambda s: pd.Series(_normalize_name_for_join(str(s))))
    M = pd.merge(L, R, on=["__ln","__fn"], how="left", suffixes=("", "_roster"))
    return M.drop(columns=["__ln","__fn"])

def _last_name_key(full_name: str) -> Tuple[str,str]:
    s = str(full_name or "").strip()
    if not s: return ("","")
    if "," in s:
        last, rest = [x.strip() for x in s.split(",",1)]
        first = rest.split(" ")[0] if rest else ""
    else:
        parts = s.split()
        first = parts[0]; last = parts[-1]
    ln = "".join(ch for ch in last.lower() if ch.isalpha())
    fn = "".join(ch for ch in first.lower() if ch.isalpha())
    return (ln, fn)

def _nz_or_blank(v: Optional[float]) -> Optional[float]:
    if v is None: return None
    try:
        f = float(v)
        return None if abs(f) < 1e-9 else round(f,3)
    except Exception:
        return None

def _norm_label(s: str) -> str:
    return "".join(ch for ch in str(s).lower() if ch.isalnum())

# ---------- Sierra detection ----------
@dataclass
class SierraLayout:
    header_row: int
    name_idx: int
    hours_idx: int
    rate_idx: Optional[int]
    days_idx: Optional[int]

def _detect_sierra_layout(df: pd.DataFrame) -> Optional[SierraLayout]:
    for r in range(min(60, len(df))):
        row = df.iloc[r].astype(str).str.strip().str.lower()
        if "name" in set(row.values) and "hours" in set(row.values):
            name_idx  = row[row=="name"].index[0]
            hours_idx = row[row=="hours"].index[0]
            rate_idx  = row[row=="rate"].index[0] if any(row=="rate") else None
            days_idx  = row[row=="days"].index[0] if any(row=="days") else None
            return SierraLayout(r,name_idx,hours_idx,rate_idx,days_idx)
    return None

# ---------- Read + aggregate ----------
def _read_sierra_records(xlsx_bytes: bytes) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns:
      agg: DF(Name, Reg_sum, OT_sum, DT_sum, Rate_last)
      per_day: DF(Name, Weekday (0=Mon..6=Sun), Hours, Rate_last)
    """
    bio = io.BytesIO(xlsx_bytes)
    df0 = pd.read_excel(bio, sheet_name=0, header=None)
    layout = _detect_sierra_layout(df0)
    if not layout:
        raise ValueError("Could not detect Sierra header row with 'Name' and 'Hours'.")

    bio.seek(0)
    df = pd.read_excel(bio, sheet_name=0, header=layout.header_row)
    df.columns = [str(c).strip() for c in df.columns]

    name_col  = df.columns[layout.name_idx]
    hours_col = df.columns[layout.hours_idx]
    df = df.rename(columns={name_col:"Name", hours_col:"Hours"})
    if layout.rate_idx is not None:
        rate_name = df.columns[layout.rate_idx]
        if rate_name not in ("Name","Hours"): df = df.rename(columns={rate_name:"Rate"})
    if layout.days_idx is not None:
        days_name = df.columns[layout.days_idx]
        df = df.rename(columns={days_name:"Days"})

    df["Name"] = df["Name"].astype(str).str.strip()
    df = df[df["Name"]!=""].copy()
    df["Hours_num"] = df["Hours"].apply(_num)
    df = df[df["Hours_num"].notnull()].copy()
    df["Rate_num"] = df["Rate"].apply(_num) if "Rate" in df.columns else None

    # --- Weekday detection: real dates OR weekday text ("Mon", "Tue", ...)
    weekday_map = {
        "monday":0, "mon":0,
        "tuesday":1, "tue":1, "tues":1,
        "wednesday":2, "wed":2,
        "thursday":3, "thu":3, "thur":3, "thurs":3,
        "friday":4, "fri":4,
        "saturday":5, "sat":5,
        "sunday":6, "sun":6
    }

    def to_weekday(val) -> Optional[int]:
        # 1) try real date
        try:
            d = pd.to_datetime(val, errors="raise")
            return int(d.weekday())
        except Exception:
            pass
        # 2) try text
        s = str(val).strip().lower()
        # keep only letters to be tolerant ("Mon.", "Mon – 9/16")
        letters = "".join(ch for ch in s if ch.isalpha())
        return weekday_map.get(letters, None)

    if "Days" in df.columns:
        df["Weekday"] = df["Days"].apply(to_weekday)
    else:
        df["Weekday"] = None  # unknown

    # Build per-day (keep Rate_last for the row)
    per_day_raw = (df.groupby(["Name","Weekday"], dropna=False)
                     .agg(Hours=("Hours_num","sum"),
                          Rate_last=("Rate_num","last"))
                     .reset_index())

    per_day = per_day_raw.copy()

    # Daily split for REG/OT/DT (used later for weekly overlay)
    def split_daily(h: float) -> Tuple[float,float,float]:
        if h is None: return (0.0,0.0,0.0)
        reg = min(h, 8.0)
        ot  = min(max(h-8.0, 0.0), 4.0)
        dt  = max(h-12.0, 0.0)
        return (reg, ot, dt)

    per_day[["Reg","OT","DT"]] = per_day["Hours"].apply(lambda h: pd.Series(split_daily(h)))

    # Aggregate by employee
    per_emp = (per_day.groupby("Name", dropna=False)
                        .agg(Reg_sum=("Reg","sum"),
                             OT_sum=("OT","sum"),
                             DT_sum=("DT","sum"),
                             Rate_last=("Rate_last","last"))
                        .reset_index())

    # Weekly overlay >40: move excess from REG to OT
    def weekly_adjust(row):
        total = float((row["Reg_sum"] or 0)+(row["OT_sum"] or 0)+(row["DT_sum"] or 0))
        if total > 40.0:
            extra = total - 40.0
            pull = min(extra, row["Reg_sum"])
            row["Reg_sum"] -= pull
            row["OT_sum"]  += pull
            extra -= pull
            if extra > 0:
                row["OT_sum"] += extra
        return row

    per_emp = per_emp.apply(weekly_adjust, axis=1)
    agg = per_emp[per_emp["Name"].notna() & (per_emp["Name"].astype(str).str.strip()!="")]
    return agg, per_day_raw  # raw hours per weekday kept for Mon–Fri HRS

# ---------- Roster (OPTIONAL) ----------
def _load_roster() -> pd.DataFrame:
    expected = ["EmpID","SSN","Employee Name","Status","Type","PayRate","Dept"]
    candidates = [
        ROSTER_PATH,
        REPO_ROOT / "roster.xlsx",
        PROJECT_ROOT / "roster.xlsx",
        HERE / "roster.xlsx",
        Path("/roster.xlsx"),
    ]
    path = next((p for p in candidates if p.exists()), None)
    if path is None:
        empty = pd.DataFrame(columns=expected)
        empty["EmpID_clean"] = pd.Series(dtype="Int64")
        empty["SSN_clean"]   = pd.Series(dtype="Int64")
        empty["EmployeeNameRoster"] = pd.Series(dtype="string")
        return empty

    roster = pd.read_excel(path, sheet_name="Roster")
    roster.columns = [str(c).strip() for c in roster.columns]
    missing = set(expected) - set(roster.columns)
    if missing:
        raise ValueError(f"Roster found at {path} but missing columns: {missing}")

    roster["EmpID_clean"] = roster["EmpID"].apply(_safe_int)
    roster["SSN_clean"]   = roster["SSN"].apply(_safe_int)
    roster["EmployeeNameRoster"] = roster["Employee Name"].astype(str)
    return roster

# ---------- assemble rows ----------
def _pad_empid(empid: Optional[int]) -> Optional[str]:
    if empid is None: return None
    s = str(empid).strip()
    try: s = str(int(float(s)))
    except Exception: return None
    return s.zfill(10)

def _calc_totals(pay_type: str, reg: float, ot: float, dt: float,
                 rate: Optional[float], default_payrate: Optional[float]) -> float:
    # Salaried → roster PayRate is period total
    if str(pay_type or "").upper().startswith("S"):
        return float(_num(default_payrate) or 0.0)
    r = _num(rate) or _num(default_payrate) or 0.0
    return float((reg * r) + (ot * 1.5 * r) + (dt * 2.0 * r))

def _build_rows(agg: pd.DataFrame, per_day_hours: pd.DataFrame, roster: pd.DataFrame) -> List[Dict[str,object]]:
    # Per-employee weekday hours pivot (0..6)
    day_pivot = (per_day_hours
                 .assign(Weekday=lambda d: d["Weekday"].apply(lambda x: int(x) if pd.notna(x) else -1))
                 .query("Weekday >= 0 and Weekday <= 6")
                 .pivot_table(index="Name", columns="Weekday", values="Hours", aggfunc="sum", fill_value=0.0))

    joined = _best_join(agg, roster, "Name", "EmployeeNameRoster")
    out: List[Dict[str,object]] = []
    for _, r in joined.iterrows():
        name_src = str(r["Name"]).strip()
        reg = float(r.get("Reg_sum",0.0) or 0.0)
        ot  = float(r.get("OT_sum",0.0) or 0.0)
        dt  = float(r.get("DT_sum",0.0) or 0.0)
        rate_si = r.get("Rate_last")

        empid = r.get("EmpID_clean"); ssn = r.get("SSN_clean")
        name_roster = r.get("EmployeeNameRoster")
        status = r.get("Status","A"); ptype = r.get("Type","H")
        payrate_roster = r.get("PayRate"); dept = r.get("Dept")

        name_final = str(name_roster).strip() if pd.notna(name_roster) and str(name_roster).strip() else name_src
        rate_final = _num(payrate_roster) if _num(payrate_roster) is not None else _num(rate_si)

        # Per-day hours (Mon..Fri indices 0..4)
        h_mon = float(day_pivot.loc[name_src, 0]) if (name_src in day_pivot.index and 0 in day_pivot.columns) else 0.0
        h_tue = float(day_pivot.loc[name_src, 1]) if (name_src in day_pivot.index and 1 in day_pivot.columns) else 0.0
        h_wed = float(day_pivot.loc[name_src, 2]) if (name_src in day_pivot.index and 2 in day_pivot.columns) else 0.0
        h_thu = float(day_pivot.loc[name_src, 3]) if (name_src in day_pivot.index and 3 in day_pivot.columns) else 0.0
        h_fri = float(day_pivot.loc[name_src, 4]) if (name_src in day_pivot.index and 4 in day_pivot.columns) else 0.0

        # Per-day totals
        r_use = _num(rate_final) or 0.0
        t_mon = h_mon * r_use
        t_tue = h_tue * r_use
        t_wed = h_wed * r_use
        t_thu = h_thu * r_use
        t_fri = h_fri * r_use

        totals_val = _calc_totals(str(ptype or "H"), reg, ot, dt, rate_final, payrate_roster)

        out.append({
            "EmpID": _pad_empid(empid),
            "SSN": ssn,
            "Employee Name": name_final,
            "Status": status if pd.notna(status) else "A",
            "Type": ptype if pd.notna(ptype) else "H",
            "Pay Rate": rate_final if rate_final is not None else "",
            "Dept": dept if pd.notna(dept) else "",
            "REGULAR": reg,
            "OVERTIME": ot,
            "DOUBLETIME": dt,
            "PC HRS MON": h_mon, "PC HRS TUE": h_tue, "PC HRS WED": h_wed, "PC HRS THU": h_thu, "PC HRS FRI": h_fri,
            "PC TTL MON": t_mon, "PC TTL TUE": t_tue, "PC TTL WED": t_wed, "PC TTL THU": t_thu, "PC TTL FRI": t_fri,
            "Totals": totals_val,
        })
    return out

# ---------- header utils ----------
def _find_columns(hdr_raw: List[str]) -> Dict[str, Optional[int]]:
    hdr_lower = [str(x or "").strip().lower() for x in hdr_raw]
    hdr_norm  = [_norm_label(x) for x in hdr_raw]

    def idx_exact(label: str) -> Optional[int]:
        lab = label.strip().lower()
        for i, txt in enumerate(hdr_lower, start=1):
            if txt == lab:
                return i
        return None
    def idx_contains(substr: str) -> Optional[int]:
        sub = substr.strip().lower()
        for i, txt in enumerate(hdr_lower, start=1):
            if txt and sub in txt:
                return i
        return None
    def idx_norm_in(aliases: List[str]) -> Optional[int]:
        alias_norm = {_norm_label(a) for a in aliases}
        for i, n in enumerate(hdr_norm, start=1):
            if n in alias_norm:
                return i
        return None

    cols = {}
    cols["payrate"] = idx_exact("pay rate") or idx_contains("pay rate") or idx_contains("rate")
    cols["reg"]     = idx_norm_in(["regular","reg","a01"])
    cols["ot"]      = idx_norm_in(["overtime","ot","a02"])
    cols["dt"]      = idx_norm_in(["doubletime","double time","dt","a03"])

    cols["empid"]  = idx_contains("emp id") or idx_exact("empid") or idx_exact("# e:26")
    cols["ssn"]    = idx_exact("ssn") or idx_contains("ssn")
    cols["name"]   = idx_exact("employee name") or idx_exact("name") or idx_contains("employee")
    cols["status"] = idx_exact("status") or idx_contains("status")
    cols["type"]   = idx_exact("type")   or idx_exact("pay type") or idx_contains("pay type") or idx_contains("type")
    cols["dept"]   = idx_exact("dept")   or idx_exact("department") or idx_contains("department") or idx_contains("dept")

    cols["hrs_mon"] = idx_norm_in(["pchrsmon","pc hrs mon"])
    cols["hrs_tue"] = idx_norm_in(["pchrstue","pc hrs tue"])
    cols["hrs_wed"] = idx_norm_in(["pchrswed","pc hrs wed"])
    cols["hrs_thu"] = idx_norm_in(["pchrsthu","pc hrs thu"])
    cols["hrs_fri"] = idx_norm_in(["pchrsfri","pc hrs fri"])

    cols["ttl_mon"] = idx_norm_in(["pcttlmon","pc ttl mon"])
    cols["ttl_tue"] = idx_norm_in(["pcttltue","pc ttl tue"])
    cols["ttl_wed"] = idx_norm_in(["pcttlwed","pc ttl wed"])
    cols["ttl_thu"] = idx_norm_in(["pcttlthu","pc ttl thu"])
    cols["ttl_fri"] = idx_norm_in(["pcttlfri","pc ttl fri"])
    return cols

# ---------- main ----------
def sierra_excel_to_wbs_bytes(xlsx_bytes: bytes) -> bytes:
    if not WBS_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"WBS template not found at {WBS_TEMPLATE_PATH}")

    agg, per_day_hours = _read_sierra_records(xlsx_bytes)
    roster = _load_roster()
    rows = _build_rows(agg, per_day_hours, roster)

    wb = load_workbook(WBS_TEMPLATE_PATH)
    ws: Worksheet = wb["WEEKLY"] if "WEEKLY" in wb.sheetnames else next((wb[s] for s in wb.sheetnames if "week" in s.lower()), wb.active)

    # Choose header row (rows 6–12; most labels)
    header_row = None; header_score = -1
    for r in range(6, 13):
        vals = [str(c.value or "").strip() for c in ws[r]]
        score = sum(1 for v in vals if v)
        if score > header_score:
            header_score = score; header_row = r
    if header_row is None: header_row = 8
    DATA_START = header_row + 1

    # Capture existing order (fixed block at top)
    hdr_raw = [str(c.value or "").strip() for c in ws[header_row]]
    cols = _find_columns(hdr_raw)
    name_col_idx = cols["name"]
    existing_order: List[str] = []
    if name_col_idx:
        for rr in range(DATA_START, ws.max_row + 1):
            val = ws.cell(row=rr, column=name_col_idx).value
            if val and str(val).strip():
                existing_order.append(str(val).strip())

    # Clear old rows
    if ws.max_row >= DATA_START:
        ws.delete_rows(DATA_START, ws.max_row - DATA_START + 1)

    last_header_col = len(hdr_raw) if hdr_raw else 20
    payrate_col = cols["payrate"] or min(last_header_col, 7)
    if not cols["reg"]: cols["reg"] = payrate_col + 1
    if not cols["ot"]:  cols["ot"]  = cols["reg"] + 1
    if not cols["dt"]:  cols["dt"]  = cols["ot"]  + 1

    # Totals candidates
    hdr_lower = [h.lower() for h in hdr_raw]
    totals_candidates: List[int] = []
    for i, txt in enumerate(hdr_lower, start=1):
        if "total" in txt and i not in totals_candidates: totals_candidates.append(i)
    right_most_labeled = max([i for i, t in enumerate(hdr_raw, start=1) if str(t).strip()], default=None)
    if right_most_labeled and right_most_labeled not in totals_candidates: totals_candidates.append(right_most_labeled)
    if last_header_col not in totals_candidates: totals_candidates.append(last_header_col)
    if (cols["dt"] + 1) not in totals_candidates: totals_candidates.append(cols["dt"] + 1)

    # Ordering: existing first, others A–Z by last name
    existing_index = {name: idx for idx, name in enumerate(existing_order)} if existing_order else {}
    def sort_key(row: Dict[str,object]) -> Tuple[int, Tuple[str,str]]:
        nm = str(row["Employee Name"] or "").strip()
        pri = existing_index.get(nm, 10**9)
        return (pri, _last_name_key(nm))

    rows_sorted = sorted(rows, key=sort_key)

    # Write rows
    r = DATA_START
    for row in rows_sorted:
        if cols["empid"]:  ws.cell(row=r, column=cols["empid"],  value=row["EmpID"])
        if cols["ssn"]:    ws.cell(row=r, column=cols["ssn"],    value=row["SSN"])
        if cols["name"]:   ws.cell(row=r, column=cols["name"],   value=row["Employee Name"])
        if cols["status"]: ws.cell(row=r, column=cols["status"], value=row["Status"])
        if cols["type"]:   ws.cell(row=r, column=cols["type"],   value=row["Type"])
        if cols["dept"]:   ws.cell(row=r, column=cols["dept"],   value=row["Dept"])

        rate_val = row["Pay Rate"] if row["Pay Rate"] != "" else None
        ws.cell(row=r, column=payrate_col, value=rate_val)
        ws.cell(row=r, column=cols["reg"], value=_nz_or_blank(row["REGULAR"]))
        ws.cell(row=r, column=cols["ot"],  value=_nz_or_blank(row["OVERTIME"]))
        ws.cell(row=r, column=cols["dt"],  value=_nz_or_blank(row["DOUBLETIME"]))

        if cols["hrs_mon"]: ws.cell(row=r, column=cols["hrs_mon"], value=_nz_or_blank(row["PC HRS MON"]))
        if cols["hrs_tue"]: ws.cell(row=r, column=cols["hrs_tue"], value=_nz_or_blank(row["PC HRS TUE"]))
        if cols["hrs_wed"]: ws.cell(row=r, column=cols["hrs_wed"], value=_nz_or_blank(row["PC HRS WED"]))
        if cols["hrs_thu"]: ws.cell(row=r, column=cols["hrs_thu"], value=_nz_or_blank(row["PC HRS THU"]))
        if cols["hrs_fri"]: ws.cell(row=r, column=cols["hrs_fri"], value=_nz_or_blank(row["PC HRS FRI"]))

        if cols["ttl_mon"]: ws.cell(row=r, column=cols["ttl_mon"], value=_nz_or_blank(row["PC TTL MON"]))
        if cols["ttl_tue"]: ws.cell(row=r, column=cols["ttl_tue"], value=_nz_or_blank(row["PC TTL TUE"]))
        if cols["ttl_wed"]: ws.cell(row=r, column=cols["ttl_wed"], value=_nz_or_blank(row["PC TTL WED"]))
        if cols["ttl_thu"]: ws.cell(row=r, column=cols["ttl_thu"], value=_nz_or_blank(row["PC TTL THU"]))
        if cols["ttl_fri"]: ws.cell(row=r, column=cols["ttl_fri"], value=_nz_or_blank(row["PC TTL FRI"]))

        rate_ref = f"{get_column_letter(payrate_col)}{r}"
        reg_ref  = f"{get_column_letter(cols['reg'])}{r}"
        ot_ref   = f"{get_column_letter(cols['ot'])}{r}"
        dt_ref   = f"{get_column_letter(cols['dt'])}{r}"
        formula = f"=({reg_ref}*{rate_ref})+({ot_ref}*1.5*{rate_ref})+({dt_ref}*2*{rate_ref})"
        for tcol in totals_candidates:
            if tcol and tcol > 0:
                ws.cell(row=r, column=tcol).value = formula

        r += 1

    # DEBUG
    if "DEBUG" in wb.sheetnames:
        wb.remove(wb["DEBUG"])
    dbg = wb.create_sheet("DEBUG")
    dbg.append([f"Header row used: {header_row}"])
    dbg.append(hdr_raw); dbg.append([])
    dbg.append([
        "Chosen columns →",
        "Pay Rate", payrate_col,
        "REG", cols["reg"], "OT", cols["ot"], "DT", cols["dt"],
        "HRS MON", cols["hrs_mon"], "HRS TUE", cols["hrs_tue"], "HRS WED", cols["hrs_wed"],
        "HRS THU", cols["hrs_thu"], "HRS FRI", cols["hrs_fri"],
        "TTL MON", cols["ttl_mon"], "TTL TUE", cols["ttl_tue"], "TTL WED", cols["ttl_wed"],
        "TTL THU", cols["ttl_thu"], "TTL FRI", cols["ttl_fri"],
    ])
    dbg.append([])
    dbg.append(["First 20 rows preview"])
    for row in rows_sorted[:20]:
        dbg.append([
            row["Employee Name"], row["Pay Rate"],
            row["REGULAR"], row["OVERTIME"], row["DOUBLETIME"],
            row["PC HRS MON"], row["PC TTL MON"],
            row["PC HRS TUE"], row["PC TTL TUE"],
            row["PC HRS WED"], row["PC TTL WED"],
            row["PC HRS THU"], row["PC TTL THU"],
            row["PC HRS FRI"], row["PC TTL FRI"],
            row["Totals"]
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()
