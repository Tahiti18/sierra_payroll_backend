# app/main.py
from __future__ import annotations

import io
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import pandas as pd
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from starlette.responses import StreamingResponse

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell

# ------------------------------------------------------------------------------
# FastAPI + CORS
# ------------------------------------------------------------------------------
app = FastAPI(title="Sierra → WBS Converter", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # tighten later if desired
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ------------------------------------------------------------------------------
# Gold-master employee order (Name, SSN) — LOCKED
# Source: “WBS Payroll 9_19_25 for Marwan 2.xlsx” (WEEKLY)
# Keep as ground truth so the order is stable week-to-week.
# ------------------------------------------------------------------------------
EMP_ORDER: List[Tuple[str, str]] = [
    ('Robleza, Dianne', '626946016'),
    ('Shafer, Emily', '622809130'),
    ('Stokes, Symone', '616259695'),
    ('Young, Giana L', '602762103'),
    ('Garcia, Bryan', '616259654'),
    ('Garcia, Miguel A', '681068099'),
    ('Hernandez, Diego', '652143527'),
    ('Pacheco Estrada, Jesus', '645935042'),
    ('Pajarito, Ramon', '685942713'),
    ('Rivas Beltran, Angel M', '358119787'),
    ('Romero Solis, Juan', '836220003'),
    ('Alcaraz, Luis', '432946242'),
    ('Alvarez, Jose', '534908967'),
    ('Arizmendi, Fernando', '613871092'),
    ('Arroyo, Jose', '364725751'),
    ('Bello, Luis', '616226754'),
    ('Bocanegra, Jose', '605908531'),
    ('Bustos, Eric', '603965173'),
    ('Cardoso, Hipolito', '264022781'),
    ('Castaneda, Andy', '611042001'),
    ('Castillo, Moises', '653246578'),
    ('Chavez, Derick J', '610591002'),
    ('Chavez, Endhy', '625379918'),
    ('Contreras, Brian', '137178003'),
    ('Cortez, Kevin', '656253574'),
    ('Cuevas, Marcelo', '625928556'),
    ('Dean, Jake', '615598883'),
    ('Duarte, Esau', '602134323'),
    ('Duarte, Kevin', '616259686'),
    ('Espinoza, Jose', '605486881'),
    ('Esquivel, Kleber', '606450473'),
    ('Flores, Saul', '625928557'),
    ('Garcia, Eduardo', '613909068'),
    ('Garcia, Miguel', '625170362'),
    ('Gomez, Jose', '897981424'),
    ('Gomez, Randel', '625928558'),
    ('Gonzalez, Alejandro', '621318203'),
    ('Gonzalez, Emanuel', '625170363'),
    ('Gonzalez, Miguel', '681068100'),
    ('Hernandez, Diego', '652143527'),
    ('Hernandez, Edgar', '625170364'),
    ('Hernandez, Sergio', '625170365'),
    ('Jose (Luis) Alvarez', '534908967'),
    ('Juan Romero Solis', '836220003'),
    ('Lopez, Alexander', '656253575'),
    ('Lopez, Yair', '625928559'),
    ('Lopez, Zefferino', '625928560'),
    ('Martinez, Alberto', '625170366'),
    ('Martinez, Emiliano', '625928561'),
    ('Martinez, Maciel', '625170367'),
    ('Moreno, Eduardo', '625928562'),
    ('Navarro, Jose', '625928563'),
    ('Pacheco, Jesus', '645935042'),
    ('Padilla, Carlos', '614425738'),
    ('Pajarito, Ramon', '685942713'),
    ('Pelagio, Miguel', '625928564'),
    ('Perez, Edgar', '625928565'),
    ('Ramon, Endhy', '625379918'),
    ('Ramos, Omar', '625928566'),
    ('Rivas, Manuel', '625928567'),
    ('Robledo, Francisco', '613108074'),
    ('Rodriguez, Anthony', '625928568'),
    ('Santos, Efrain', '625928569'),
    ('Santos, Javier', '625928570'),
    ('Solis, Juan Romero', '836220003'),
    ('Stokes, Symone', '616259695'),
    ('Torres, Anthony', '625928571'),
    ('Torrez, Jose', '625928572'),
    ('Vera, Erick', '625928573'),
    ('Vera, Victor', '625928574'),
    ('Zamara, Cesar', '625483371'),
    ('Anolin, Robert M', '552251095'),
    ('Dean, Joe P', '556534609'),
    ('Garrido, Raul', '657554426'),
    ('Magallanes, Julio', '612219002'),
    ('Padilla, Alex', '569697404'),
    ('Pealatere, Francis', '625098739'),
    ('Phein, Saeng Tsing', '624722627'),
    ('Rios, Jose D', '530358447'),
    ('Gomez, Jose', '897981424'),
    ('Nava, Juan M', '636667958'),
    ('Padilla, Carlos', '614425738'),
    ('Robledo, Francisco', '613108074'),
]

# Sierra → bucket aliases (flexible header matching)
BUCKET_ALIASES: Dict[str, List[str]] = {
    "REG": ["A01", "REG", "REGULAR"],
    "OT": ["A02", "OT", "OVERTIME"],
    "DT": ["A03", "DT", "DOUBLETIME", "DOUBLE TIME"],
    "VAC": ["A06", "VAC", "VACATION"],
    "SICK": ["A07", "SICK"],
    "HOLIDAY": ["A08", "HOLIDAY"],
    "BONUS": ["A04", "BONUS"],
    "COMM": ["A05", "COMM", "COMMISSION"],
    # Optional extras
    "TRAVEL": ["ATE", "TRAVEL"],
    "PC_HRS_MON": ["AH1"],
    "PC_TTL_MON": ["AI1"],
    "PC_HRS_TUE": ["AH2"],
    "PC_TTL_TUE": ["AI2"],
    "PC_HRS_WED": ["AH3"],
    "PC_TTL_WED": ["AI3"],
    "PC_HRS_THU": ["AH4"],
    "PC_TTL_THU": ["AI4"],
    "PC_HRS_FRI": ["AH5"],
    "PC_TTL_FRI": ["AI5"],
}

# ------------------------------------------------------------------------------
# Small utils
# ------------------------------------------------------------------------------
def _norm(s: str) -> str:
    return str(s or "").strip().lower().replace("_", "").replace(" ", "")

def _clean_ssn(s: str) -> str:
    return "".join(ch for ch in str(s or "") if ch.isdigit())

def _first_nonempty(vals: List[object]) -> str:
    for v in vals:
        if v is None:
            continue
        t = str(v).strip()
        if t and t.lower() != "nan":
            return t
    return ""

# ------------------------------------------------------------------------------
# Sierra reader — expects the *weekly summary* style (REG/OT/DT/etc. columns)
# It auto-finds the header row and flex-maps the columns.
# ------------------------------------------------------------------------------
def read_sierra_weekly(io_bytes: bytes) -> pd.DataFrame:
    try:
        xl = pd.ExcelFile(io.BytesIO(io_bytes))
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Cannot open Excel: {e}")

    sheet = "WEEKLY" if "WEEKLY" in xl.sheet_names else xl.sheet_names[0]
    raw = pd.read_excel(io.BytesIO(io_bytes), sheet_name=sheet, header=None)

    # Find header block: a row that contains both "employee" and "ssn"
    header_row = None
    for i in range(min(60, len(raw))):
        row = [_norm(x) for x in raw.iloc[i].tolist()]
        if any("employee" in x for x in row) and any("ssn" in x for x in row):
            header_row = i
            break
    if header_row is None:
        raise HTTPException(status_code=422, detail="Could not locate header row (need 'Employee' and 'SSN').")

    # Many exports have a descriptor line + a code line -> use the *next* row as headers
    df1 = pd.read_excel(io.BytesIO(io_bytes), sheet_name=sheet, header=header_row + 1)
    # Promote first data row as final headers (common Sierra quirk)
    df1.columns = df1.iloc[0]
    df1 = df1.iloc[1:].reset_index(drop=True)

    # Build normalized name → actual column map
    name_map: Dict[str, str] = {}
    for c in df1.columns:
        name_map[_norm(str(c))] = str(c)

    def pick(*cands: str) -> Optional[str]:
        for c in cands:
            n = _norm(c)
            if n in name_map:
                return name_map[n]
            # relaxed contains
            for k, v in name_map.items():
                if n in k or k in n:
                    return v
        return None

    col_emp = pick("Employee Name", "Employee")
    col_ssn = pick("SSN", "Social")
    col_status = pick("Status")
    col_type = pick("Type")
    col_rate = pick("Pay Rate", "Rate", "Hourly Rate")
    col_dept = pick("Dept", "Department")

    if not col_emp or not col_ssn:
        raise HTTPException(status_code=422, detail="Missing Employee or SSN column.")

    # Map buckets
    bucket_cols: Dict[str, Optional[str]] = {}
    for key, aliases in BUCKET_ALIASES.items():
        bucket_cols[key] = pick(*aliases)

    # Select relevant columns
    sel = [c for c in [col_emp, col_ssn, col_status, col_type, col_rate, col_dept] if c]
    for v in bucket_cols.values():
        if v and v not in sel:
            sel.append(v)

    df = df1[sel].copy()

    # Clean types
    df[col_emp] = df[col_emp].astype(str).str.strip()
    df[col_ssn] = df[col_ssn].map(_clean_ssn)

    # Make numeric buckets numeric
    for v in bucket_cols.values():
        if v:
            df[v] = pd.to_numeric(df[v], errors="coerce").fillna(0.0)

    # Aggregate duplicates (by SSN; fallback to name)
    key = df[col_ssn]
    key = key.where(key.str.len() > 0, df[col_emp])
    df["_key"] = key

    def first(series: pd.Series) -> str:
        return _first_nonempty(series.tolist())

    agg_map: Dict[str, str] = {}
    # numeric sums
    for v in bucket_cols.values():
        if v:
            agg_map[v] = "sum"
    # identity
    agg_map[col_emp] = first
    agg_map[col_ssn] = first
    if col_status: agg_map[col_status] = first
    if col_type: agg_map[col_type] = first
    if col_rate: agg_map[col_rate] = first
    if col_dept: agg_map[col_dept] = first

    g = df.groupby("_key", dropna=False).agg(agg_map).reset_index(drop=True)

    # Normalize to canonical keys
    rows: List[Dict[str, object]] = []
    for _, r in g.iterrows():
        out: Dict[str, object] = {
            "EMP_NAME": r.get(col_emp, ""),
            "SSN": r.get(col_ssn, ""),
            "STATUS": r.get(col_status, ""),
            "TYPE": r.get(col_type, ""),
            "PAY_RATE": r.get(col_rate, ""),
            "DEPT": r.get(col_dept, ""),
        }
        for k, v in bucket_cols.items():
            out[k] = float(r.get(v, 0.0)) if v else 0.0
        rows.append(out)

    return pd.DataFrame(rows)

# ------------------------------------------------------------------------------
# Write into template with header-driven column detection
# ------------------------------------------------------------------------------
def detect_wbs_columns(ws: Worksheet, scan_rows: int = 25) -> Dict[str, int]:
    """
    Scan the first N rows to find the header row that includes 'employee' and 'ssn',
    then build a 1-based column index map for all needed fields by header text.
    Works even if Status/SSN/Type moved.
    """
    header_row_idx = None
    for r in range(1, scan_rows + 1):
        row_vals = [str(ws.cell(row=r, column=c).value or "") for c in range(1, ws.max_column + 1)]
        row_norm = [_norm(x) for x in row_vals]
        if any("employee" in x for x in row_norm) and any("ssn" in x for x in row_norm):
            header_row_idx = r
            break
    if header_row_idx is None:
        # Fallback to a common default header row
        header_row_idx = 8  # typical in your template
    # Map header text → column index
    head_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        text = str(ws.cell(row=header_row_idx, column=c).value or "")
        head_map[_norm(text)] = c

    def find_col(*labels: str, contains: List[str] | None = None, default: Optional[int] = None) -> Optional[int]:
        # exact match on any candidate
        for lab in labels:
            key = _norm(lab)
            if key in head_map:
                return head_map[key]
        # contains search
        if contains:
            for k, idx in head_map.items():
                if any(_norm(x) in k for x in contains):
                    return idx
        return default

    colmap: Dict[str, int] = {}
    # Identity columns (text)
    colmap["EMP_NAME"] = find_col("Employee", contains=["employee"])
    colmap["SSN"] = find_col("SSN", contains=["ssn"])
    colmap["STATUS"] = find_col("Status", contains=["status"])
    colmap["TYPE"] = find_col("Type", contains=["type"])
    colmap["PAY_RATE"] = find_col("Pay Rate", contains=["payrate","pay","rate"])
    colmap["DEPT"] = find_col("Dept", contains=["dept","department"])

    # Buckets (hours/amounts)
    colmap["REG"] = find_col("REG (A01)", contains=["a01","reg"])
    colmap["OT"] = find_col("OT (A02)", contains=["a02","ot","overtime"])
    colmap["DT"] = find_col("DT (A03)", contains=["a03","dt","double"])
    colmap["VAC"] = find_col("VACATION (A06)", contains=["a06","vac"])
    colmap["SICK"] = find_col("SICK (A07)", contains=["a07","sick"])
    colmap["HOLIDAY"] = find_col("HOLIDAY (A08)", contains=["a08","holiday"])
    colmap["BONUS"] = find_col("BONUS (A04)", contains=["a04","bonus"])
    colmap["COMM"] = find_col("COMMISSION (A05)", contains=["a05","comm"])

    # Piece rows, travel, comments (if present)
    colmap["PC_HRS_MON"] = find_col("AH1", contains=["ah1"])
    colmap["PC_TTL_MON"] = find_col("AI1", contains=["ai1"])
    colmap["PC_HRS_TUE"] = find_col("AH2", contains=["ah2"])
    colmap["PC_TTL_TUE"] = find_col("AI2", contains=["ai2"])
    colmap["PC_HRS_WED"] = find_col("AH3", contains=["ah3"])
    colmap["PC_TTL_WED"] = find_col("AI3", contains=["ai3"])
    colmap["PC_HRS_THU"] = find_col("AH4", contains=["ah4"])
    colmap["PC_TTL_THU"] = find_col("AI4", contains=["ai4"])
    colmap["PC_HRS_FRI"] = find_col("AH5", contains=["ah5"])
    colmap["PC_TTL_FRI"] = find_col("AI5", contains=["ai5"])
    colmap["TRAVEL"] = find_col("ATE", contains=["ate","travel"])
    colmap["COMMENTS"] = find_col("Comments", contains=["comment","memo","notes"])

    # Totals column (keep formulas)
    colmap["TOTALS"] = find_col("Totals", contains=["total"])

    # Data start row is header_row + 1 (skip header)
    data_start = header_row_idx + 1
    return {"__DATA_START__": data_start, **{k: v for k, v in colmap.items() if v}}

def clear_existing_rows(ws: Worksheet, first_data_row: int, last_col: int):
    """Clear prior data rows while keeping formulas/borders (skip merged/read-only)."""
    max_row = ws.max_row or first_data_row
    for r in range(first_data_row, max_row + 1):
        # If whole data area empty already, skip
        try:
            if all((ws.cell(row=r, column=c).value in (None, "")) for c in range(1, last_col + 1)):
                continue
        except Exception:
            pass
        for c in range(1, last_col + 1):
            try:
                cell = ws.cell(row=r, column=c)
                _ = cell.value  # trigger read-only if merged
                cell.value = None
            except Exception:
                continue  # merged / read-only → leave as-is

def sort_to_locked_order(df: pd.DataFrame) -> List[Dict[str, object]]:
    by_ssn = {str(r.get("SSN", "")).strip(): r for _, r in df.iterrows()}
    by_name = {str(r.get("EMP_NAME", "")).strip(): r for _, r in df.iterrows()}
    out: List[Dict[str, object]] = []
    seen: set = set()

    # 1) in locked order
    for name, ssn in EMP_ORDER:
        r = None
        if ssn and ssn in by_ssn:
            r = by_ssn[ssn]
        elif name in by_name:
            r = by_name[name]
        if r is not None:
            key = (r.get("SSN") or "") or (r.get("EMP_NAME") or "")
            if key not in seen:
                out.append(r)
                seen.add(key)

    # 2) any extras appended (stable by name)
    for _, r in df.sort_values(by=["EMP_NAME"]).iterrows():
        key = (r.get("SSN") or "") or (r.get("EMP_NAME") or "")
        if key not in seen:
            out.append(r)
            seen.add(key)

    return out

def write_to_template(template_path: Path, rows: List[Dict[str, object]]) -> bytes:
    wb = load_workbook(str(template_path))
    sheet_name = "WEEKLY"
    if sheet_name not in wb.sheetnames:
        raise HTTPException(status_code=500, detail=f"Template missing sheet '{sheet_name}'")
    ws = wb[sheet_name]

    colmap = detect_wbs_columns(ws)
    data_start = colmap.pop("__DATA_START__")
    last_col = max(colmap.values()) if colmap else ws.max_column

    # Clear old data
    clear_existing_rows(ws, data_start, last_col)

    # Write new
    r = data_start
    for row in rows:
        # Identity fields
        def set_txt(key: str, value: object):
            if key in colmap and colmap[key]:
                ws.cell(row=r, column=colmap[key]).value = "" if value is None else str(value)

        def set_num(key: str, value: object):
            if key in colmap and colmap[key]:
                try:
                    v = float(value or 0)
                    ws.cell(row=r, column=colmap[key]).value = (None if v == 0 else v)
                except Exception:
                    ws.cell(row=r, column=colmap[key]).value = None

        set_txt("EMP_NAME", row.get("EMP_NAME", ""))
        set_txt("SSN", (row.get("SSN") and _clean_ssn(row.get("SSN"))) or "")
        set_txt("STATUS", row.get("STATUS", ""))
        set_txt("TYPE", row.get("TYPE", ""))
        set_txt("DEPT", row.get("DEPT", ""))
        # Pay rate might be text or number in Sierra export
        rate = row.get("PAY_RATE", "")
        try:
            rate_num = float(str(rate).replace(",", ""))
        except Exception:
            rate_num = None
        if "PAY_RATE" in colmap and colmap["PAY_RATE"]:
            ws.cell(row=r, column=colmap["PAY_RATE"]).value = rate_num if rate_num not in (None, 0) else None

        # Buckets
        for k in [
            "REG","OT","DT","VAC","SICK","HOLIDAY","BONUS","COMM",
            "PC_HRS_MON","PC_TTL_MON","PC_HRS_TUE","PC_TTL_TUE",
            "PC_HRS_WED","PC_TTL_WED","PC_HRS_THU","PC_TTL_THU",
            "PC_HRS_FRI","PC_TTL_FRI","TRAVEL"
        ]:
            set_num(k, row.get(k, 0))

        # DO NOT touch "TOTALS" column—template formulas remain intact
        r += 1

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

# ------------------------------------------------------------------------------
# Routes
# ------------------------------------------------------------------------------
@app.get("/health")
def health():
    return {"ok": True}

@app.post("/process-payroll")
async def process_payroll(file: UploadFile = File(...)):
    name = (file.filename or "").lower()
    if not name.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=415, detail="Please upload an Excel file (.xlsx or .xls).")

    data = await file.read()

    # 1) Parse Sierra (weekly summary)
    df = read_sierra_weekly(data)  # raises 422 if headers missing

    # 2) Lock ordering to gold master
    ordered = sort_to_locked_order(df)

    # 3) Load template from repo root and write rows
    template_path = Path(__file__).resolve().parent.parent / "wbs_template.xlsx"
    if not template_path.exists():
        raise HTTPException(status_code=500, detail=f"Template not found at {template_path}")

    out_bytes = write_to_template(template_path, ordered)

    # 4) Respond with XLSX
    out_name = f"WBS_{Path(file.filename).stem}.xlsx"
    return StreamingResponse(
        io.BytesIO(out_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'}
    )
