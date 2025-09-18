# app/converter.py
import io
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# -------------------------
# Constants / column map
# -------------------------
WBS_SHEET_NAME = "WEEKLY"          # your template tab name
WBS_DATA_START_ROW = 8             # first data row underneath headers

# Target columns (1-based indexes in the WBS template)
COL = {
    "SSN": 1,
    "EMP_NAME": 2,
    "STATUS": 3,           # A/H flag goes into Type; Status column is "A" etc.
    "TYPE": 4,             # H or S
    "PAY_RATE": 5,
    "DEPT": 6,

    "REG": 7,              # A01
    "OT": 8,               # A02
    "DT": 9,               # A03

    "VAC": 10,             # A06
    "SICK": 11,            # A07
    "HOL": 12,             # A08

    "BONUS": 13,           # A04
    "COMM": 14,            # A05

    # Piecework blocks – hours and totals Mon..Fri
    "PC_HRS_MON": 15,      # AH1
    "PC_TTL_MON": 16,      # AI1
    "PC_HRS_TUE": 17,      # AH2
    "PC_TTL_TUE": 18,      # AI2
    "PC_HRS_WED": 19,      # AH3
    "PC_TTL_WED": 20,      # AI3
    "PC_HRS_THU": 21,      # AH4
    "PC_TTL_THU": 22,      # AI4
    "PC_HRS_FRI": 23,      # AH5
    "PC_TTL_FRI": 24,      # AI5

    "TRAVEL": 25,          # ATE
    "NOTES": 26,           # Comments
    "TOTALS": 27,          # pink totals column at far right
}

ALLOWED_INPUT_EXTS = (".xlsx", ".xls")


# -------------------------
# Helpers
# -------------------------
def _std(s: str) -> str:
    return (s or "").strip().lower().replace("\n", " ").replace("\r", " ")


def _to_date(x) -> Optional[date]:
    if pd.isna(x):
        return None
    if isinstance(x, datetime):
        return x.date()
    try:
        return pd.to_datetime(x).date()
    except Exception:
        return None


def _normalize_name(raw: str) -> str:
    if not isinstance(raw, str):
        return ""
    s = raw.strip()
    if not s:
        return s
    # keep the exact “Last, First” if already in that shape; otherwise do "Last, First"
    parts = [p for p in s.replace(",", " ").split() if p]
    if len(parts) >= 2:
        first = parts[0]
        last = parts[-1]
        return f"{last}, {first}"
    return s


def _apply_ca_daily_ot(day_hours: float) -> Tuple[float, float, float]:
    """Return (REG, OT, DT) from a single day of hours using CA daily split."""
    h = float(day_hours or 0.0)
    reg = min(h, 8.0)
    ot = 0.0
    dt = 0.0
    if h > 8:
        ot = min(h - 8.0, 4.0)
    if h > 12:
        dt = h - 12.0
    return reg, ot, dt


def _load_roster(repo_root: Path) -> pd.DataFrame:
    """
    Load roster.xlsx (preferred) or roster.csv from the repo root.
    Expected headers (case/spacing flexible): name, ssn, status, type, rate, dept, order
    """
    for fname in ["roster.xlsx", "roster.xls", "roster.csv"]:
        p = repo_root / fname
        if p.exists():
            if p.suffix.lower() == ".csv":
                df = pd.read_csv(p)
            else:
                df = pd.read_excel(p, engine="openpyxl")
            break
    else:
        # No roster file found — return empty
        return pd.DataFrame(columns=["name", "ssn", "status", "type", "rate", "dept", "order"])

    cols = {_std(c): c for c in df.columns}
    def pick(*cands):
        for c in cands:
            cc = cols.get(_std(c))
            if cc:
                return df[cc]
        return pd.Series([None] * len(df))

    out = pd.DataFrame({
        "name":  pick("name", "employee", "employee name").apply(lambda s: _normalize_name(str(s) if pd.notna(s) else "")),
        "ssn":   pick("ssn", "social", "social security", "social security number"),
        "status":pick("status"),
        "type":  pick("type", "pay type", "employee type"),
        "rate":  pd.to_numeric(pick("rate", "pay rate", "hourly rate", "wage"), errors="coerce"),
        "dept":  pick("dept", "department", "division"),
        "order": pd.to_numeric(pick("order", "sort", "rank"), errors="coerce"),
    })
    return out


def _safe_set(ws: Worksheet, row: int, col: int, value):
    """
    Write to (row,col) if it isn't a merged read-only cell.
    If the target cell is in a merged range, write to the top-left anchor of that range.
    """
    cell = ws.cell(row=row, column=col)
    # openpyxl marks merged followers with MergedCell type
    if cell.__class__.__name__ == "MergedCell":
        # find merged range that contains this coordinate
        for mr in ws.merged_cells.ranges:
            if (row, col) in mr:
                anchor = ws.cell(row=mr.min_row, column=mr.min_col)
                anchor.value = value
                return
        # if not found (shouldn't happen), just skip
        return
    cell.value = value


def _money(x) -> float:
    return float(0.0 if x is None or (isinstance(x, float) and pd.isna(x)) else x)


# -------------------------
# Core conversion function
# -------------------------
def convert_sierra_to_wbs(input_bytes: bytes, sheet_name: Optional[str] = None) -> bytes:
    """
    Convert a Sierra weekly file to WBS format using the wbs_template.xlsx in repo root
    and roster-based ordering/identity.
    """
    # 1) Load Sierra input
    excel = pd.ExcelFile(io.BytesIO(input_bytes))
    src_sheet = sheet_name or excel.sheet_names[0]
    df = excel.parse(src_sheet)

    if df.empty:
        raise ValueError("Input Excel appears to be empty.")

    # Soft-match columns
    def find_col(cands: List[str]) -> Optional[str]:
        cols = {_std(c): c for c in df.columns}
        for want in cands:
            k = _std(want)
            if k in cols:
                return cols[k]
        # relaxed contains
        for want in cands:
            k = _std(want)
            for std_name, real in cols.items():
                if k in std_name:
                    return real
        return None

    req_map = {
        "employee": ["employee", "employee name", "name", "worker"],
        "date":     ["date", "work date", "worked date", "day"],
        "hours":    ["hours", "hrs", "total hours", "work hours"],
        "rate":     ["rate", "pay rate", "hourly rate", "wage"],
    }
    opt_map = {
        "task":     ["task", "earn type", "earning", "code"],
        "dept":     ["dept", "department", "division"],
        "bonus":    ["bonus"],
        "comm":     ["commission", "comm"],
        "travel":   ["travel amount", "travel", "mileage"],
        "notes":    ["notes", "comments", "remark"],
        "pc_mon":   ["pc hrs mon", "piece hrs mon", "ah1", "pc mon"],
        "pc_tmon":  ["pc ttl mon", "piece ttl mon", "ai1"],
        "pc_tue":   ["pc hrs tue", "ah2"],
        "pc_ttue":  ["pc ttl tue", "ai2"],
        "pc_wed":   ["pc hrs wed", "ah3"],
        "pc_twed":  ["pc ttl wed", "ai3"],
        "pc_thu":   ["pc hrs thu", "ah4"],
        "pc_tthu":  ["pc ttl thu", "ai4"],
        "pc_fri":   ["pc hrs fri", "ah5"],
        "pc_tfri":  ["pc ttl fri", "ai5"],
        "vac":      ["vac", "vacation"],
        "sick":     ["sick"],
        "hol":      ["holiday", "hol"],
    }

    resolved_req: Dict[str, str] = {}
    missing = []
    for key, cands in req_map.items():
        col = find_col(cands)
        if not col:
            missing.append(f"{key} (any of: {', '.join(cands)})")
        else:
            resolved_req[key] = col
    if missing:
        raise ValueError("Missing required columns: " + "; ".join(missing))

    resolved_opt = {k: find_col(v) for k, v in opt_map.items()}

    core = df[[resolved_req["employee"], resolved_req["date"], resolved_req["hours"], resolved_req["rate"]]].copy()
    core.columns = ["employee", "date", "hours", "rate"]

    # attach selected extras
    for k, col in resolved_opt.items():
        core[k] = df[col] if col else None

    # normalize
    core["employee"] = core["employee"].astype(str).map(_normalize_name)
    core["date"] = core["date"].map(_to_date)
    core["hours"] = pd.to_numeric(core["hours"], errors="coerce").fillna(0.0).astype(float)
    core["rate"] = pd.to_numeric(core["rate"], errors="coerce")

    core = core[(core["employee"].str.len() > 0) & core["date"].notna() & (core["hours"] > 0)]

    # 2) Daily split then weekly sums per (employee, rate)
    # Group by employee+date to apply daily OT/DT
    daily = core.groupby(["employee", "date", "rate"], dropna=False)["hours"].sum().reset_index()
    split_rows = []
    for _, r in daily.iterrows():
        reg, ot, dt = _apply_ca_daily_ot(float(r["hours"]))
        split_rows.append({
            "employee": r["employee"],
            "rate": float(r["rate"]) if pd.notna(r["rate"]) else None,
            "REG": reg, "OT": ot, "DT": dt
        })
    split_df = pd.DataFrame(split_rows)

    weekly = split_df.groupby(["employee", "rate"], dropna=False)[["REG", "OT", "DT"]].sum().reset_index()

    # Carry “piecework” / add-ons by summing per employee if those columns exist
    add_cols = ["vac", "sick", "hol", "bonus", "comm", "travel",
                "pc_mon", "pc_tmon", "pc_tue", "pc_ttue", "pc_wed", "pc_twed",
                "pc_thu", "pc_tthu", "pc_fri", "pc_tfri", "notes"]
    agg_map = {c: "sum" for c in add_cols if c != "notes"}
    if "notes" in add_cols:
        agg_map["notes"] = "first"

    adds = (core.groupby("employee", dropna=False)[[c for c in add_cols if c in core.columns]]
            .agg(agg_map) if add_cols else pd.DataFrame({"employee": []}))
    if not adds.empty:
        weekly = weekly.merge(adds.reset_index(), on="employee", how="left")

    # 3) Roster lookups: SSN / Status / Type / Dept / PayRate / fixed order
    here = Path(__file__).resolve().parents[1]  # repo root (two up from app/)
    roster = _load_roster(here)

    if not roster.empty:
        weekly = weekly.merge(
            roster.rename(columns={
                "name": "employee",
                "ssn": "SSN",
                "status": "Status",
                "type": "Type",
                "rate": "RosterRate",
                "dept": "Dept",
                "order": "Order"
            }),
            on="employee", how="left"
        )
        # Prefer roster rate if present
        weekly["FinalRate"] = weekly["RosterRate"].combine_first(weekly["rate"])
    else:
        weekly["SSN"] = ""
        weekly["Status"] = "A"
        weekly["Type"] = "H"
        weekly["Dept"] = ""
        weekly["Order"] = pd.NA
        weekly["FinalRate"] = weekly["rate"]

    # default Type normalization (H/S)
    weekly["Type"] = weekly["Type"].astype(str).str.upper().map(lambda x: "S" if x.startswith("S") else "H")
    weekly["Status"] = weekly["Status"].fillna("A")

    # 4) Compute totals in dollars (used for the pink 'Totals' if your template isn't already a formula)
    rate = weekly["FinalRate"].fillna(0.0).astype(float)
    weekly["TOT_$"] = (
        weekly["REG"].fillna(0) * rate +
        weekly["OT"].fillna(0)  * rate * 1.5 +
        weekly["DT"].fillna(0)  * rate * 2.0 +
        _money(0.0)  # placeholder to keep style
    )
    # Add-ons in hours/dollars
    for c in ["vac", "sick", "hol"]:
        if c in weekly.columns:
            weekly["TOT_$"] += weekly[c].fillna(0) * rate
    for c in ["bonus", "comm", "travel"]:
        if c in weekly.columns:
            weekly["TOT_$"] += weekly[c].fillna(0)

    # 5) Sort: by explicit roster Order (ascending), then keep stable by roster appearance then name
    weekly["Order"] = pd.to_numeric(weekly["Order"], errors="coerce")
    weekly = weekly.sort_values(by=["Order", "employee"], kind="stable", na_position="last")

    # 6) Open template and write
    template_path = here / "wbs_template.xlsx"
    if not template_path.exists():
        raise ValueError(f"WBS template not found at {template_path}")

    wb = load_workbook(template_path)
    if WBS_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Template sheet '{WBS_SHEET_NAME}' not found in template.")
    ws = wb[WBS_SHEET_NAME]

    # Clear previous data area (values only, keep styles & merged headers)
    max_row = ws.max_row
    if max_row >= WBS_DATA_START_ROW:
        for r in range(WBS_DATA_START_ROW, max_row + 1):
            # Clear only data columns (1..COL["TOTALS"])
            for c in range(1, COL["TOTALS"] + 1):
                cell = ws.cell(row=r, column=c)
                if cell.__class__.__name__ == "MergedCell":
                    # never touch merged header followers
                    continue
                cell.value = None

    # Write rows
    row_cursor = WBS_DATA_START_ROW
    def num_get(df, col):
        return df[col] if col in df.columns else 0.0

    for _, r in weekly.iterrows():
        _safe_set(ws, row_cursor, COL["SSN"],     r.get("SSN", ""))
        _safe_set(ws, row_cursor, COL["EMP_NAME"], r["employee"])
        _safe_set(ws, row_cursor, COL["STATUS"],  r.get("Status", "A"))
        _safe_set(ws, row_cursor, COL["TYPE"],    r.get("Type", "H"))
        _safe_set(ws, row_cursor, COL["PAY_RATE"], round(_money(r.get("FinalRate", 0.0)), 2))
        _safe_set(ws, row_cursor, COL["DEPT"],    r.get("Dept", ""))

        _safe_set(ws, row_cursor, COL["REG"], round(_money(r.get("REG", 0.0)), 3))
        _safe_set(ws, row_cursor, COL["OT"],  round(_money(r.get("OT", 0.0)), 3))
        _safe_set(ws, row_cursor, COL["DT"],  round(_money(r.get("DT", 0.0)), 3))

        _safe_set(ws, row_cursor, COL["VAC"], round(_money(r.get("vac", 0.0)), 3))
        _safe_set(ws, row_cursor, COL["SICK"],round(_money(r.get("sick", 0.0)), 3))
        _safe_set(ws, row_cursor, COL["HOL"], round(_money(r.get("hol", 0.0)), 3))

        _safe_set(ws, row_cursor, COL["BONUS"], round(_money(r.get("bonus", 0.0)), 2))
        _safe_set(ws, row_cursor, COL["COMM"],  round(_money(r.get("comm", 0.0)), 2))

        _safe_set(ws, row_cursor, COL["PC_HRS_MON"], round(_money(r.get("pc_mon", 0.0)), 3))
        _safe_set(ws, row_cursor, COL["PC_TTL_MON"], round(_money(r.get("pc_tmon", 0.0)), 2))
        _safe_set(ws, row_cursor, COL["PC_HRS_TUE"], round(_money(r.get("pc_tue", 0.0)), 3))
        _safe_set(ws, row_cursor, COL["PC_TTL_TUE"], round(_money(r.get("pc_ttue", 0.0)), 2))
        _safe_set(ws, row_cursor, COL["PC_HRS_WED"], round(_money(r.get("pc_wed", 0.0)), 3))
        _safe_set(ws, row_cursor, COL["PC_TTL_WED"], round(_money(r.get("pc_twed", 0.0)), 2))
        _safe_set(ws, row_cursor, COL["PC_HRS_THU"], round(_money(r.get("pc_thu", 0.0)), 3))
        _safe_set(ws, row_cursor, COL["PC_TTL_THU"], round(_money(r.get("pc_tthu", 0.0)), 2))
        _safe_set(ws, row_cursor, COL["PC_HRS_FRI"], round(_money(r.get("pc_fri", 0.0)), 3))
        _safe_set(ws, row_cursor, COL["PC_TTL_FRI"], round(_money(r.get("pc_tfri", 0.0)), 2))

        _safe_set(ws, row_cursor, COL["TRAVEL"], round(_money(r.get("travel", 0.0)), 2))
        _safe_set(ws, row_cursor, COL["NOTES"],  r.get("notes", ""))

        # If your template already has a formula in the Totals column, DO NOT overwrite it.
        # If it doesn't, write our computed dollars:
        cell_tot = ws.cell(row=row_cursor, column=COL["TOTALS"])
        if cell_tot.value in (None, ""):
            _safe_set(ws, row_cursor, COL["TOTALS"], round(_money(r.get("TOT_$", 0.0)), 2))

        row_cursor += 1

    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()
